import xml.etree.ElementTree as ET

import openpyxl
from config import (BASE_TAG, ENCODING, FILENAME, OUT_FILENAME_XML,
                    SUBBASE_TAG, TEMP_XML)
from tqdm import tqdm

XML_DEC = True

if not ENCODING:
    XML_DEC = False
    print('Дефолтная кодировка')

def csv_from_excel(filename):
    wb = openpyxl.load_workbook(filename)
    sh = wb.active
    return [row for row in tqdm(sh.iter_rows(max_col=sh.max_column, max_row=sh.max_row,values_only=True),
                                                                  desc=f'Собираем данные из {filename}')]


def convert_to_xml():
    data_list = csv_from_excel(FILENAME)
    fields =  data_list[0:1]
    root = ET.Element(BASE_TAG)
    for i in tqdm(range(1, len(data_list)), desc='Запаковываем в бл#$ский xml'): # перебираем списки двумерного массива
        doc = ET.SubElement(root, SUBBASE_TAG)
        for j in fields[0]: # перебираем файл с именами полей
            __iter_elem=fields[0].index(j)
            if j in TEMP_XML.keys():
                for tag, attr in TEMP_XML[j].items(): # находим по ключу поле
                    ET.SubElement(doc, tag, attr).text=str(data_list[i][__iter_elem]) # находим список по итератору i находим элемент итерации __iter_elem
    tree = ET.ElementTree(root)
    tree.write(OUT_FILENAME_XML, encoding = ENCODING, xml_declaration = XML_DEC)
    print(f'Конвертация прошла успешно! файл: {FILENAME}')


if __name__ == '__main__':
    convert_to_xml()
